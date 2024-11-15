import openpyxl
import openpyxl.workbook
from examples.faker import data_samples

def create():
    book = openpyxl.Workbook()
    print(book.sheetnames)
    book.remove(book.active)
    book.create_sheet("Коллеги", 0)
    book.create_sheet("Черный список")
    book.create_sheet("Клиенты")
    
    
    for sheet in book.worksheets:
        for row in data_samples():
            sheet.append(row)
    book.save("Test.xlsx")


