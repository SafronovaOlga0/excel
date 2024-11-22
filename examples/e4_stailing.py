import openpyxl
from openpyxl.styles import Font, PatternFill, NamedStyle

def stailing():
    book = openpyxl.load_workbook("Test.xlsx")
    sheet = book["Коллеги"]
    sheet.column_dimensions["A"].width = 100
    sheet.column_dimensions["B"].width = 50


    sheet.column_dimensions["A"].font = Font(size = 24, color='0070C0')
    font1 = Font(b=True, size=23, color='002455')
    fill1=PatternFill(fill_type='solid', fgColor='FFFF00')
    sheet['A1'].font=font1
    sheet['B1'].font=font1

    # sheet['A1'].style="Good"

    lgustyle=NamedStyle(name="lgustyle")
    lgustyle.font=font1
    lgustyle.fill=fill1

    sheet.cell(row=1, column=1).style=lgustyle


    book.save("Test.xlsx")